from django.shortcuts import render, get_object_or_404
from django.http import HttpResponse
from django.db.models import Sum, Count, Avg
from django.utils import timezone
import datetime
# Modelos
# pyrefly: ignore [missing-import]
from .models import Jugador, Libro, Rompecabezas
# pyrefly: ignore [missing-import]
from .models import HistorialLectura, HistorialAhorcado, HistorialRompecabezas, HistorialTrofeo
# pyrefly: ignore [missing-import]
from .views import admin_required

# ReportLab para PDF
from reportlab.lib.pagesizes import letter
from reportlab.lib import colors
from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment

import os
from django.conf import settings

def dibujar_encabezado(canvas, doc):
    canvas.saveState()
    logo_path = os.path.join(settings.BASE_DIR, 'aplicacion', 'static', 'inicio_sesion', 'img', 'logolecto.png')
    
    if os.path.exists(logo_path):
        try:
            from PIL import Image
            from reportlab.lib.utils import ImageReader
            # Procesamos la imagen para asegurarnos de que el fondo transparente se vuelva blanco
            # ya que ReportLab a veces renderiza la transparencia como negro.
            with Image.open(logo_path) as img:
                img = img.convert("RGBA")
                background = Image.new('RGBA', img.size, (255, 255, 255, 255))
                background.paste(img, mask=img)
                rgb_img = background.convert('RGB')
                logo_to_draw = ImageReader(rgb_img)
        except Exception:
            # Fallback a la ruta directa si PIL falla
            logo_to_draw = logo_path
            
        # Encabezado con logo de Lectosoft
        # Es necesario especificar tanto width como height para que preserveAspectRatio funcione correctamente
        canvas.drawImage(logo_to_draw, 40, 730, width=150, height=45, preserveAspectRatio=True)
        
        # Línea decorativa debajo del encabezado
        canvas.setStrokeColor(colors.HexColor('#4F81BD'))
        canvas.setLineWidth(2)
        canvas.line(40, 715, 570, 715)
        
        # --- PIE DE PÁGINA (FOOTER) ---
        canvas.setStrokeColor(colors.HexColor('#4F81BD'))
        canvas.setLineWidth(1)
        canvas.line(40, 50, 570, 50)
        
        canvas.setFont("Helvetica", 9)
        canvas.setFillColor(colors.HexColor('#666666'))
        
        # Texto de contacto
        footer_text = "Lectosoft - Plataforma de Lectura Dinámica | Contacto: info@lectosoft.com | Sede: Medellín, Colombia"
        # Centrar texto
        canvas.drawCentredString(306, 35, footer_text)
        
    canvas.restoreState()

@admin_required
def reportes_generales(request):
    total_estudiantes = Jugador.objects.count()
    total_libros_leidos = HistorialLectura.objects.count()
    total_rompecabezas = HistorialRompecabezas.objects.count()
    total_ahorcado = HistorialAhorcado.objects.count()
    total_trofeos = HistorialTrofeo.objects.count()

    # Ranking (Top 3)
    ranking_puntos = Jugador.objects.all().order_by('-puntuacion')[:3]
    
    # Graficos stats
    promedio_comprension = HistorialLectura.objects.aggregate(Avg('porcentaje_comprension'))['porcentaje_comprension__avg'] or 0

    # Estadísticas Avanzadas de Rompecabezas
    historial_romp = HistorialRompecabezas.objects.all()
    total_partidas_romp = historial_romp.count()
    tiempo_total_romp = 0
    vel_total_romp = 0
    prec_total_romp = 0
    rompecabezas_count = {}
    
    for h in historial_romp:
        minutos = (h.tiempo_empleado / 60) if h.tiempo_empleado > 0 else (1/60)
        piezas = h.dificultad * h.dificultad
        velocidad = piezas / minutos
        precision = 0
        if h.movimientos_totales > 0:
            precision = (h.movimientos_correctos / h.movimientos_totales) * 100
        tiempo_total_romp += h.tiempo_empleado
        vel_total_romp += velocidad
        prec_total_romp += precision
        titulo = h.rompecabezas.titulo
        rompecabezas_count[titulo] = rompecabezas_count.get(titulo, 0) + 1
        
    tiempo_prom_romp = (tiempo_total_romp / total_partidas_romp) if total_partidas_romp > 0 else 0
    vel_prom_romp = (vel_total_romp / total_partidas_romp) if total_partidas_romp > 0 else 0
    prec_prom_romp = (prec_total_romp / total_partidas_romp) if total_partidas_romp > 0 else 0
    
    romp_mas_jugado = max(rompecabezas_count, key=rompecabezas_count.get) if rompecabezas_count else "N/A"
    
    mejores_tiempos_romp = historial_romp.order_by('tiempo_empleado')[:3]

    return render(request, 'reportes_generales.html', {
        'total_estudiantes': total_estudiantes,
        'total_libros_leidos': total_libros_leidos,
        'total_rompecabezas': total_rompecabezas,
        'total_ahorcado': total_ahorcado,
        'total_trofeos': total_trofeos,
        'ranking_puntos': ranking_puntos,
        'promedio_comprension': round(promedio_comprension, 2),
        'promedio_comprension_entero': int(promedio_comprension),
        'promedio_comprension_str': f"{promedio_comprension:.1f}".replace(',', '.'),
        
        # Variables de rompecabezas
        'tiempo_prom_romp': round(tiempo_prom_romp, 1),
        'vel_prom_romp': round(vel_prom_romp, 1),
        'prec_prom_romp': round(prec_prom_romp, 1),
        'romp_mas_jugado': romp_mas_jugado,
        'mejores_tiempos_romp': mejores_tiempos_romp,
    })

@admin_required
def reporte_jugador_pdf(request, jugador_id):
    jugador = get_object_or_404(Jugador, id=jugador_id)
    response = HttpResponse(content_type='application/pdf')
    response['Content-Disposition'] = f'attachment; filename="reporte_{jugador.nombre_usuario}.pdf"'

    doc = SimpleDocTemplate(response, pagesize=letter, topMargin=95, bottomMargin=70)
    elements = []
    styles = getSampleStyleSheet()
    title_style = styles['Heading1']
    subtitle_style = styles['Heading2']
    normal_style = styles['Normal']

    # Título
    elements.append(Paragraph(f"Reporte Individual: {jugador.nombre_usuario}", title_style))
    elements.append(Spacer(1, 12))

    # Información General
    elements.append(Paragraph("Información General", subtitle_style))
    elements.append(Paragraph(f"<b>Fecha de registro:</b> {jugador.fecha_creacion.strftime('%d/%m/%Y')}", normal_style))
    elements.append(Paragraph(f"<b>Puntos Acumulados:</b> {jugador.puntuacion}", normal_style))
    elements.append(Paragraph(f"<b>Fichas Obtenidas:</b> {jugador.fichas_ganadas}", normal_style))
    trofeos_totales = jugador.historial_trofeos.count()
    elements.append(Paragraph(f"<b>Trofeos Totales:</b> {trofeos_totales}", normal_style))
    elements.append(Spacer(1, 12))

    # Lectura
    elements.append(Paragraph("Estadísticas de Lectura", subtitle_style))
    lecturas = jugador.historial_lecturas.all()
    if lecturas.exists():
        data_lectura = [["Cuento", "Pág. Leídas", "Preguntas", "Correctas", "% Comprensión"]]
        for lec in lecturas:
            data_lectura.append([
                lec.libro.titulo,
                str(lec.paginas_leidas),
                str(lec.preguntas_totales),
                str(lec.preguntas_correctas),
                f"{lec.porcentaje_comprension}%"
            ])
        t_lectura = Table(data_lectura, style=[
            ('BACKGROUND', (0,0), (-1,0), colors.grey),
            ('TEXTCOLOR', (0,0), (-1,0), colors.whitesmoke),
            ('ALIGN', (0,0), (-1,-1), 'CENTER'),
            ('GRID', (0,0), (-1,-1), 1, colors.black)
        ])
        elements.append(t_lectura)
    else:
        elements.append(Paragraph("No hay registros de lectura.", normal_style))
    elements.append(Spacer(1, 12))

    # Ahorcado
    elements.append(Paragraph("Rendimiento Ahorcado", subtitle_style))
    ahorcados = jugador.historial_ahorcado.all()
    if ahorcados.exists():
        data_ahorcado = [["Palabra", "Tiempo (s)", "Intentos", "Resultado"]]
        for ah in ahorcados:
            res = "Ganó" if ah.resultado else "Perdió"
            data_ahorcado.append([ah.palabra, f"{ah.tiempo_respuesta}s", str(ah.intentos), res])
        t_ahorcado = Table(data_ahorcado, style=[
            ('BACKGROUND', (0,0), (-1,0), colors.grey),
            ('TEXTCOLOR', (0,0), (-1,0), colors.whitesmoke),
            ('ALIGN', (0,0), (-1,-1), 'CENTER'),
            ('GRID', (0,0), (-1,-1), 1, colors.black)
        ])
        elements.append(t_ahorcado)
    else:
        elements.append(Paragraph("No hay registros de ahorcado.", normal_style))
    elements.append(Spacer(1, 12))

    # Rompecabezas
    elements.append(Paragraph("Rendimiento Rompecabezas", subtitle_style))
    rompecabezas = jugador.historial_rompecabezas.all().order_by('fecha')
    if rompecabezas.exists():
        data_romp = [["Fecha", "Rompecabezas", "Tiempo", "Vel.", "Prec.", "Err."]]
        for rom in rompecabezas:
            minutos = (rom.tiempo_empleado / 60) if rom.tiempo_empleado > 0 else (1/60)
            piezas = rom.dificultad * rom.dificultad
            vel = round(piezas / minutos, 1)
            prec = 0
            if rom.movimientos_totales > 0:
                prec = round((rom.movimientos_correctos / rom.movimientos_totales) * 100, 1)
            err = max(0, rom.movimientos_totales - rom.movimientos_correctos)
            
            data_romp.append([
                rom.fecha.strftime('%d/%m/%Y'),
                rom.rompecabezas.titulo[:15],
                f"{round(rom.tiempo_empleado, 1)}s",
                f"{vel} p/m",
                f"{prec}%",
                str(err)
            ])
        t_romp = Table(data_romp, style=[
            ('BACKGROUND', (0,0), (-1,0), colors.grey),
            ('TEXTCOLOR', (0,0), (-1,0), colors.whitesmoke),
            ('ALIGN', (0,0), (-1,-1), 'CENTER'),
            ('GRID', (0,0), (-1,-1), 1, colors.black),
            ('FONTNAME', (0,0), (-1,0), 'Helvetica-Bold'),
        ])
        elements.append(t_romp)
        
        # Observacion especifica de rompecabezas
        prec_prom = sum([float(x[4].replace('%','')) for x in data_romp[1:]]) / len(data_romp[1:])
        if prec_prom >= 80:
            obs_romp = "¡Excelente habilidad espacial! Gran precisión."
        elif prec_prom >= 50:
            obs_romp = "Buen progreso. Puede mejorar intentando cometer menos errores."
        else:
            obs_romp = "Se requiere más práctica para mejorar la atención visual."
        elements.append(Spacer(1, 5))
        elements.append(Paragraph(f"<b>Obs. Rompecabezas:</b> {obs_romp}", normal_style))
        
    else:
        elements.append(Paragraph("No hay registros de rompecabezas.", normal_style))
    elements.append(Spacer(1, 12))

    # Observación Automática
    elements.append(Paragraph("Observación", subtitle_style))
    promedio_comp = lecturas.aggregate(Avg('porcentaje_comprension'))['porcentaje_comprension__avg'] or 0
    if promedio_comp >= 80:
        obs = "El estudiante demuestra una excelente comprensión lectora y participación activa."
    elif promedio_comp >= 50:
        obs = "El estudiante tiene un nivel promedio. Se recomienda fomentar más la lectura para mejorar la retención."
    elif lecturas.count() == 0:
        obs = "El estudiante aún no ha registrado lecturas. Motívelo a comenzar su aventura."
    else:
        obs = "El estudiante está presentando dificultades en la comprensión lectora. Se sugiere acompañamiento."
    elements.append(Paragraph(obs, normal_style))

    doc.build(elements, onFirstPage=dibujar_encabezado, onLaterPages=dibujar_encabezado)
    return response

@admin_required
def exportar_reporte_excel(request):
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename="reporte_general_lectosoft.xlsx"'

    wb = Workbook()
    
    # Estilos comunes
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
    
    def apply_header_style(ws, headers):
        ws.append(headers)
        for cell in ws[1]:
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center")

    # Hoja 1: Estudiantes
    ws1 = wb.active
    ws1.title = "Estudiantes"
    apply_header_style(ws1, ["Nombre", "Puntos", "Fichas", "Libros Leídos", "Trofeos", "Comprensión Promedio"])
    for j in Jugador.objects.all():
        libros = j.historial_lecturas.count()
        trofeos = j.historial_trofeos.count()
        comp = j.historial_lecturas.aggregate(Avg('porcentaje_comprension'))['porcentaje_comprension__avg'] or 0
        ws1.append([j.nombre_usuario, j.puntuacion, j.fichas_ganadas, libros, trofeos, round(comp, 2)])

    # Hoja 2: Lecturas
    ws2 = wb.create_sheet(title="Lecturas")
    apply_header_style(ws2, ["Libro", "Veces Leído", "Promedio Comprensión"])
    for lib in Libro.objects.all():
        veces = HistorialLectura.objects.filter(libro=lib).count()
        comp = HistorialLectura.objects.filter(libro=lib).aggregate(Avg('porcentaje_comprension'))['porcentaje_comprension__avg'] or 0
        ws2.append([lib.titulo, veces, round(comp, 2)])

    # Hoja 3: Ahorcado
    ws3 = wb.create_sheet(title="Ahorcado")
    apply_header_style(ws3, ["Estudiante", "Partidas Jugadas", "Aciertos", "Tiempo Promedio"])
    for j in Jugador.objects.all():
        partidas = j.historial_ahorcado.count()
        aciertos = j.historial_ahorcado.filter(resultado=True).count()
        tiempo = j.historial_ahorcado.aggregate(Avg('tiempo_respuesta'))['tiempo_respuesta__avg'] or 0
        if partidas > 0:
            ws3.append([j.nombre_usuario, partidas, aciertos, round(tiempo, 2)])

    # Hoja 4: Rompecabezas
    ws4 = wb.create_sheet(title="Rompecabezas")
    apply_header_style(ws4, ["Estudiante", "Rompecabezas Completados", "Tiempo Promedio (s)", "Velocidad Promedio (piezas/min)", "Precisión Promedio (%)"])
    for j in Jugador.objects.all():
        romps = j.historial_rompecabezas.all()
        completados = romps.count()
        if completados > 0:
            tiempo = romps.aggregate(Avg('tiempo_empleado'))['tiempo_empleado__avg'] or 0
            
            # Calcular velocidad y precision
            vel_total = 0
            prec_total = 0
            for r in romps:
                minutos = (r.tiempo_empleado / 60) if r.tiempo_empleado > 0 else 1
                piezas = r.dificultad * r.dificultad
                vel_total += (piezas / minutos)
                if r.movimientos_totales > 0:
                    prec_total += (r.movimientos_correctos / r.movimientos_totales) * 100
                else:
                    prec_total += 0
            vel_prom = vel_total / completados
            prec_prom = prec_total / completados
            
            ws4.append([j.nombre_usuario, completados, round(tiempo, 2), round(vel_prom, 2), round(prec_prom, 2)])

    # Hoja 5: Trofeos
    ws5 = wb.create_sheet(title="Trofeos")
    apply_header_style(ws5, ["Estudiante", "Bronce", "Plata", "Oro"])
    for j in Jugador.objects.all():
        bronce = j.historial_trofeos.filter(tipo_trofeo='BRONCE').count()
        plata = j.historial_trofeos.filter(tipo_trofeo='PLATA').count()
        oro = j.historial_trofeos.filter(tipo_trofeo='ORO').count()
        if (bronce + plata + oro) > 0:
            ws5.append([j.nombre_usuario, bronce, plata, oro])

    wb.save(response)
    return response

# --- NUEVAS VISTAS PARA ROMPECABEZAS ---

def reporte_individual_rompecabezas(request, jugador_id):
    jugador = get_object_or_404(Jugador, id=jugador_id)
    # Validacion: solo el mismo jugador o admin puede verlo (opcional, aqui dejamos pasar para que sea facil de probar)
    
    historial = jugador.historial_rompecabezas.all().order_by('fecha')
    
    # Procesar historial para los graficos y stats
    datos_historial = []
    tiempo_total = 0
    vel_total = 0
    prec_total = 0
    mejor_tiempo = float('inf')
    
    for h in historial:
        minutos = (h.tiempo_empleado / 60) if h.tiempo_empleado > 0 else (1/60) # evitar division por 0
        piezas = h.dificultad * h.dificultad
        velocidad = piezas / minutos
        
        precision = 0
        if h.movimientos_totales > 0:
            precision = (h.movimientos_correctos / h.movimientos_totales) * 100
            
        if h.tiempo_empleado < mejor_tiempo:
            mejor_tiempo = h.tiempo_empleado
            
        tiempo_total += h.tiempo_empleado
        vel_total += velocidad
        prec_total += precision
        
        datos_historial.append({
            'fecha': h.fecha.strftime('%d/%m/%Y %H:%M'),
            'titulo': h.rompecabezas.titulo,
            'piezas': piezas,
            'tiempo': round(h.tiempo_empleado, 1),
            'velocidad': round(velocidad, 2),
            'precision': round(precision, 1),
            'clase_precision': 'color-mejor' if precision > 80 else 'color-peor',
            'errores': max(0, h.movimientos_totales - h.movimientos_correctos)
        })
        
    completados = len(datos_historial)
    tiempo_promedio = (tiempo_total / completados) if completados > 0 else 0
    vel_promedio = (vel_total / completados) if completados > 0 else 0
    prec_promedio = (prec_total / completados) if completados > 0 else 0
    
    if mejor_tiempo == float('inf'):
        mejor_tiempo = 0

    # Determinar si mejoró o empeoró comparando la ultima y penultima
    estado_mejora = "Mantuvo"
    clase_mejora = "color-igual"
    if completados >= 2:
        ult = datos_historial[-1]
        penult = datos_historial[-2]
        if ult['velocidad'] > penult['velocidad'] or ult['precision'] > penult['precision']:
            estado_mejora = "Mejoró"
            clase_mejora = "color-mejor"
        elif ult['velocidad'] < penult['velocidad'] or ult['precision'] < penult['precision']:
            estado_mejora = "Empeoró"
            clase_mejora = "color-peor"

    import json
    context = {
        'jugador': jugador,
        'datos_historial': list(reversed(datos_historial)), # Mostrar mas reciente primero en la tabla
        'grafico_datos': json.dumps(datos_historial), # Cronologico para grafico
        'completados': completados,
        'tiempo_promedio': round(tiempo_promedio, 1),
        'vel_promedio': round(vel_promedio, 1),
        'prec_promedio': round(prec_promedio, 1),
        'mejor_tiempo': round(mejor_tiempo, 1),
        'estado_mejora': estado_mejora,
        'clase_mejora': clase_mejora,
    }
    
    return render(request, 'reporte_individual_rompecabezas.html', context)

@admin_required
def estadisticas_rompecabezas_admin(request):
    historial = HistorialRompecabezas.objects.all()
    total_partidas = historial.count()
    total_ninos = historial.values('jugador').distinct().count()
    
    tiempo_total = 0
    vel_total = 0
    prec_total = 0
    
    rompecabezas_count = {}
    
    for h in historial:
        minutos = (h.tiempo_empleado / 60) if h.tiempo_empleado > 0 else (1/60)
        piezas = h.dificultad * h.dificultad
        velocidad = piezas / minutos
        
        precision = 0
        if h.movimientos_totales > 0:
            precision = (h.movimientos_correctos / h.movimientos_totales) * 100
            
        tiempo_total += h.tiempo_empleado
        vel_total += velocidad
        prec_total += precision
        
        titulo = h.rompecabezas.titulo
        rompecabezas_count[titulo] = rompecabezas_count.get(titulo, 0) + 1
        
    tiempo_prom = (tiempo_total / total_partidas) if total_partidas > 0 else 0
    vel_prom = (vel_total / total_partidas) if total_partidas > 0 else 0
    prec_prom = (prec_total / total_partidas) if total_partidas > 0 else 0
    
    romp_mas_jugado = max(rompecabezas_count, key=rompecabezas_count.get) if rompecabezas_count else "N/A"
    romp_menos_jugado = min(rompecabezas_count, key=rompecabezas_count.get) if rompecabezas_count else "N/A"

    # Preparar el ranking de mejores tiempos
    # Vamos a obtener los 5 con menor tiempo, pero asegurando que la precision sea decente o simplemente menor tiempo
    mejores_tiempos = historial.order_by('tiempo_empleado')[:5]

    context = {
        'total_ninos': total_ninos,
        'total_partidas': total_partidas,
        'tiempo_prom': round(tiempo_prom, 1),
        'vel_prom': round(vel_prom, 1),
        'prec_prom': round(prec_prom, 1),
        'romp_mas_jugado': romp_mas_jugado,
        'romp_menos_jugado': romp_menos_jugado,
        'mejores_tiempos': mejores_tiempos,
    }
    
    return render(request, 'reportes_rompecabezas_general.html', context)

def exportar_pdf_rompecabezas_kid(request, jugador_id):
    # Genera un PDF parecido a reporte_jugador_pdf pero enfocado solo en rompecabezas
    jugador = get_object_or_404(Jugador, id=jugador_id)
    response = HttpResponse(content_type='application/pdf')
    response['Content-Disposition'] = f'attachment; filename="reporte_rompecabezas_{jugador.nombre_usuario}.pdf"'

    doc = SimpleDocTemplate(response, pagesize=letter, topMargin=95, bottomMargin=70)
    elements = []
    styles = getSampleStyleSheet()
    
    elements.append(Paragraph(f"Reporte de Rompecabezas: {jugador.nombre_usuario}", styles['Heading1']))
    elements.append(Spacer(1, 12))

    historial = jugador.historial_rompecabezas.all().order_by('fecha')
    if not historial.exists():
        elements.append(Paragraph("Aún no ha jugado ningún rompecabezas.", styles['Normal']))
        doc.build(elements, onFirstPage=dibujar_encabezado, onLaterPages=dibujar_encabezado)
        return response

    data_romp = [["Fecha", "Rompecabezas", "Tiempo", "Velocidad", "Precisión", "Errores"]]
    
    for h in historial:
        minutos = (h.tiempo_empleado / 60) if h.tiempo_empleado > 0 else (1/60)
        piezas = h.dificultad * h.dificultad
        vel = round(piezas / minutos, 1)
        prec = 0
        if h.movimientos_totales > 0:
            prec = round((h.movimientos_correctos / h.movimientos_totales) * 100, 1)
        err = max(0, h.movimientos_totales - h.movimientos_correctos)
        
        data_romp.append([
            h.fecha.strftime('%d/%m/%Y'),
            h.rompecabezas.titulo[:20],
            f"{round(h.tiempo_empleado, 1)}s",
            f"{vel} p/m",
            f"{prec}%",
            str(err)
        ])
        
    t_romp = Table(data_romp, style=[
        ('BACKGROUND', (0,0), (-1,0), colors.HexColor('#FFB6C1')),
        ('TEXTCOLOR', (0,0), (-1,0), colors.black),
        ('ALIGN', (0,0), (-1,-1), 'CENTER'),
        ('GRID', (0,0), (-1,-1), 1, colors.black),
        ('FONTNAME', (0,0), (-1,0), 'Helvetica-Bold'),
    ])
    elements.append(t_romp)
    
    # Observacion
    elements.append(Spacer(1, 20))
    prec_prom = sum([float(x[4].replace('%','')) for x in data_romp[1:]]) / len(data_romp[1:])
    if prec_prom >= 80:
        obs = "¡Excelente habilidad espacial y paciencia! El niño tiene gran precisión."
    elif prec_prom >= 50:
        obs = "Buen progreso. Puede mejorar intentando cometer menos errores antes de colocar la pieza."
    else:
        obs = "Se requiere más práctica para mejorar la atención visual y la colocación de piezas."
    
    elements.append(Paragraph(f"<b>Observación:</b> {obs}", styles['Normal']))
    
    doc.build(elements, onFirstPage=dibujar_encabezado, onLaterPages=dibujar_encabezado)
    return response

@admin_required
def exportar_reporte_pdf(request):
    mes = request.GET.get('mes')
    response = HttpResponse(content_type='application/pdf')
    
    filename = f'reporte_general_lectosoft_mes_{mes}.pdf' if mes and mes.isdigit() else f'reporte_general_{datetime.date.today()}.pdf'
    response['Content-Disposition'] = f'attachment; filename="{filename}"'

    doc = SimpleDocTemplate(response, pagesize=letter, topMargin=95, bottomMargin=70)
    elements = []
    styles = getSampleStyleSheet()
    title_style = styles['Heading1']
    subtitle_style = styles['Heading2']
    normal_style = styles['Normal']

    elements.append(Paragraph("Reporte General Consolidado - Lectosoft", title_style))
    elements.append(Spacer(1, 12))

    # Título
    titulo_texto = f"Reporte General: LectoSoft (Mes {mes})" if mes and mes.isdigit() else "Reporte General: LectoSoft"
    elements.append(Paragraph(titulo_texto, title_style))
    elements.append(Spacer(1, 12))

    total_estudiantes = Jugador.objects.count()
    
    if mes and mes.isdigit():
        mes_num = int(mes)
        total_libros_leidos = HistorialLectura.objects.filter(fecha_inicio__month=mes_num).count()
        total_rompecabezas = HistorialRompecabezas.objects.filter(fecha__month=mes_num).count()
        total_ahorcado = HistorialAhorcado.objects.filter(fecha__month=mes_num).count()
        total_trofeos = HistorialTrofeo.objects.filter(fecha_obtenida__month=mes_num).count()
        promedio_comprension = HistorialLectura.objects.filter(fecha_inicio__month=mes_num).aggregate(Avg('porcentaje_comprension'))['porcentaje_comprension__avg'] or 0
    else:
        total_libros_leidos = HistorialLectura.objects.count()
        total_rompecabezas = HistorialRompecabezas.objects.count()
        total_ahorcado = HistorialAhorcado.objects.count()
        total_trofeos = HistorialTrofeo.objects.count()
        promedio_comprension = HistorialLectura.objects.aggregate(Avg('porcentaje_comprension'))['porcentaje_comprension__avg'] or 0

    elements.append(Paragraph("Resumen de Estadísticas", subtitle_style))
    data_resumen = [
        ["Estudiantes", "Libros Leídos", "Rompecabezas", "Ahorcado", "Trofeos", "Comprensión"],
        [str(total_estudiantes), str(total_libros_leidos), str(total_rompecabezas), str(total_ahorcado), str(total_trofeos), f"{round(promedio_comprension, 2)}%"]
    ]
    t_resumen = Table(data_resumen, style=[
        ('BACKGROUND', (0,0), (-1,0), colors.grey),
        ('TEXTCOLOR', (0,0), (-1,0), colors.whitesmoke),
        ('ALIGN', (0,0), (-1,-1), 'CENTER'),
        ('GRID', (0,0), (-1,-1), 1, colors.black)
    ])
    elements.append(t_resumen)
    elements.append(Spacer(1, 20))

    # Ranking
    elements.append(Paragraph("Top 10 Puntuaciones Globales", subtitle_style))
    ranking = Jugador.objects.all().order_by('-puntuacion')[:10]
    if ranking.exists():
        data_ranking = [["Posición", "Jugador", "Puntuación"]]
        for idx, jug in enumerate(ranking):
            data_ranking.append([f"#{idx+1}", jug.nombre_usuario, f"{jug.puntuacion} pts"])
        t_ranking = Table(data_ranking, style=[
            ('BACKGROUND', (0,0), (-1,0), colors.grey),
            ('TEXTCOLOR', (0,0), (-1,0), colors.whitesmoke),
            ('ALIGN', (0,0), (-1,-1), 'CENTER'),
            ('GRID', (0,0), (-1,-1), 1, colors.black)
        ])
        elements.append(t_ranking)
    else:
        elements.append(Paragraph("No hay suficientes jugadores.", normal_style))
        
    doc.build(elements, onFirstPage=dibujar_encabezado, onLaterPages=dibujar_encabezado)
    return response
